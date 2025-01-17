"""
EMIT: Classify interference type GUI
----------------------------------------
This example uses a GUI to open an AEDT project with
an EMIT design and analyze the results to classify the
worst-case interference. 
"""
###############################################################################
# Perform required imports
# ~~~~~~~~~~~~~~~~~~~~~~~~
# Perform required imports.

import sys
from pyaedt.emit_core.emit_constants import InterfererType, ResultType, TxRxMode
from pyaedt import Emit
import pyaedt
import os
import subprocess
import pyaedt.generic.constants as consts

# Check that emit is a compatible version
emitapp_desktop_version = "2023.2"
if emitapp_desktop_version < "2023.2":
    print("Must have v2023.2 or later")
    sys.exit()

# Check to see which Python libraries have been installed
reqs = subprocess.check_output([sys.executable, '-m', 'pip', 'freeze'])
installed_packages = [r.decode().split('==')[0] for r in reqs.split()]

# Install required packages if they are not installed
def install(package):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])

# Install required libraries for GUI and Excel exporting (internet connection needed)
required_packages = ['PySide6', 'openpyxl']
for package in required_packages:
    if package not in installed_packages:
        install(package)

# Import PySide6 and openpyxl libraries
from PySide6 import QtWidgets, QtUiTools, QtGui
from openpyxl.styles import PatternFill
import openpyxl

# Uncomment if there are Qt plugin errors
# import PySide6
# dirname = os.path.dirname(PySide6.__file__)
# plugin_path = os.path.join(dirname, 'plugins', 'platforms')
# os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path

# Launch EMIT
non_graphical = False
new_thread = True
d = pyaedt.launch_desktop(emitapp_desktop_version, non_graphical, new_thread)
emitapp = Emit(pyaedt.generate_unique_project_name())

# Add emitapi to system path
emit_path = emitapp._desktop_install_dir + "/Delcross"
sys.path.append(emit_path)
import EmitApiPython
api = EmitApiPython.EmitApi()

# Define .ui file for GUI
ui_file = pyaedt.downloads.download_file("emit", "interference_gui.ui")
Ui_MainWindow, _ = QtUiTools.loadUiType(ui_file)

class DoubleDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, decimals):
        super().__init__()
        self.decimals = decimals

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QtWidgets.QLineEdit):
            validator = QtGui.QDoubleValidator(parent)
            validator.setDecimals(self.decimals)
            editor.setValidator(validator)
        return editor

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.setup_widgets()

    ###############################################################################
    # Setup widgets
    # ~~~~~~~~~~~~~
    # Define all widgets from the UI file, connect the widgets to functions, define 
    # table colors, and format table settings. 
    
    def setup_widgets(self):
        # Widget definitions for file selection/tab management
        self.file_select_btn = self.findChild(QtWidgets.QToolButton, "file_select_btn")
        self.file_path_box = self.findChild(QtWidgets.QLineEdit, "file_path_box") 
        self.design_name_dropdown = self.findChild(QtWidgets.QComboBox, "design_name_dropdown")
        self.tab_widget = self.findChild(QtWidgets.QTabWidget, "tab_widget")

        # Widget definitions for protection level classification
        self.protection_results_btn = self.findChild(QtWidgets.QPushButton, "protection_results_btn")
        self.protection_matrix = self.findChild(QtWidgets.QTableWidget, "protection_matrix")    
        self.protection_legend_table = self.findChild(QtWidgets.QTableWidget, "protection_legend_table")
        self.damage_check = self.findChild(QtWidgets.QCheckBox, "damage_check")
        self.overload_check = self.findChild(QtWidgets.QCheckBox, "overload_check")
        self.intermodulation_check = self.findChild(QtWidgets.QCheckBox, "intermodulation_check")
        self.desensitization_check = self.findChild(QtWidgets.QCheckBox, "desensitization_check")
        self.protection_export_btn = self.findChild(QtWidgets.QPushButton, "protection_export_btn")
        self.radio_specific_levels = self.findChild(QtWidgets.QCheckBox, "radio_specific_levels")
        self.radio_dropdown = self.findChild(QtWidgets.QComboBox, "radio_dropdown")
        self.protection_save_img_btn = self.findChild(QtWidgets.QPushButton, 'protection_save_img_btn')

        # Setup for protection level buttons and table
        self.protection_results_btn.setEnabled(False)
        self.protection_export_btn.setEnabled(False)
        self.protection_save_img_btn.setEnabled(False)
        self.file_select_btn.clicked.connect(self.open_file_dialog) 
        self.protection_export_btn.clicked.connect(self.save_results_excel)
        self.protection_results_btn.clicked.connect(self.protection_results)
        self.protection_legend_table.resizeRowsToContents()
        self.protection_legend_table.resizeColumnsToContents()
        self.damage_check.stateChanged.connect(self.protection_results)
        self.overload_check.stateChanged.connect(self.protection_results)
        self.intermodulation_check.stateChanged.connect(self.protection_results)
        self.desensitization_check.stateChanged.connect(self.protection_results)
        self.protection_legend_table.setEditTriggers(QtWidgets.QTableWidget.DoubleClicked)
        self.global_protection_level = True
        self.protection_levels = {}
        values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
        self.protection_levels['Global'] = values
        self.changing = False
        self.radio_dropdown.currentIndexChanged.connect(self.radio_dropdown_changed)
        self.protection_legend_table.itemChanged.connect(self.table_changed)
        self.protection_save_img_btn.clicked.connect(self.save_image)

        # Widget definitions for interference type
        self.interference_results_btn = self.findChild(QtWidgets.QPushButton, "interference_results_btn")
        self.interference_matrix = self.findChild(QtWidgets.QTableWidget, "interference_matrix")    
        self.interference_legend_table = self.findChild(QtWidgets.QTableWidget, "interference_legend_table")
        self.in_in_check = self.findChild(QtWidgets.QCheckBox, "in_in_check")
        self.in_out_check = self.findChild(QtWidgets.QCheckBox, "in_out_check")
        self.out_in_check = self.findChild(QtWidgets.QCheckBox, "out_in_check")
        self.out_out_check = self.findChild(QtWidgets.QCheckBox, "out_out_check")
        self.interference_export_btn = self.findChild(QtWidgets.QPushButton, "interference_export_btn")
        self.interference_save_img_btn = self.findChild(QtWidgets.QPushButton, 'interference_save_img_btn')
        
        # Setup for interference type buttons and table
        self.interference_results_btn.setEnabled(False)
        self.interference_export_btn.setEnabled(False)
        self.interference_save_img_btn.setEnabled(False)
        self.interference_export_btn.clicked.connect(self.save_results_excel)
        self.interference_results_btn.clicked.connect(self.interference_results)
        self.interference_legend_table.resizeRowsToContents() 
        self.interference_legend_table.resizeColumnsToContents()
        self.in_in_check.stateChanged.connect(self.interference_results)
        self.in_out_check.stateChanged.connect(self.interference_results)
        self.out_in_check.stateChanged.connect(self.interference_results)
        self.out_out_check.stateChanged.connect(self.interference_results)
        self.radio_specific_levels.stateChanged.connect(self.radio_specific)
        self.interference_save_img_btn.clicked.connect(self.save_image)

        # Color definition dictionary and previous project/design names
        self.color_dict = {"green": [QtGui.QColor(125, 115, 202),'#7d73ca'], 
                           "yellow":[QtGui.QColor(211, 89, 162), '#d359a2'], 
                           "orange": [QtGui.QColor(255, 99, 97), '#ff6361'], 
                           "red": [QtGui.QColor(255, 166, 0), '#ffa600'], 
                           "white": [QtGui.QColor("white"),'#ffffff']}
        self.previous_design = ''
        self.previous_project = ''

        # Set the legend tables to strech resize mode
        header = self.protection_legend_table.horizontalHeader()
        v_header = self.protection_legend_table.verticalHeader()

        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch)

        header = self.interference_legend_table.horizontalHeader()
        v_header = self.interference_legend_table.verticalHeader()

        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch)

        # Input validation for protection level legend table
        delegate = DoubleDelegate(decimals=2)
        self.protection_legend_table.setItemDelegateForColumn(0, delegate)
        self.open_file_dialog()

    ###############################################################################
    # Open file dialog and select project
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Open the file dialog for project selection and populate the design dropdown 
    # with all EMIT designs in the project. 

    def open_file_dialog(self):
        fname = QtWidgets.QFileDialog.getOpenFileName(self, "Select EMIT Project", "", "Ansys Electronics Desktop Files (*.aedt)", )
        if fname: 
            self.file_path_box.setText(fname[0])

            # Close previous project and open specified one
            emitapp.close_project()       
            emitapp.load_project(self.file_path_box.text())

            # Check if project is already open
            if emitapp.lock_file == None:
                msg = QtWidgets.QMessageBox()
                msg.setWindowTitle("Error: Project already open")
                msg.setText("Project is locked. Close or remove the lock before proceeding. See AEDT log for more information.")
                x = msg.exec()
                return
            
            # Populate design dropdown with all design names
            designs = emitapp.oproject.GetDesigns()
            emit_designs = [d for d in designs if d.GetDesignType() == "EMIT"]
            self.design_name_dropdown.clear()
            for d in emit_designs:
                self.design_name_dropdown.addItem(d.GetName())
                
            self.design_name_dropdown.setEnabled(True)

            if self.radio_specific_levels.isEnabled():
                self.radio_specific_levels.setChecked(False)
                self.radio_dropdown.clear()
                self.radio_dropdown.setEnabled(False)
                self.protection_levels = {}
                values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
                self.protection_levels['Global'] = values

            self.radio_specific_levels.setEnabled(True)
            self.protection_results_btn.setEnabled(True)
            self.interference_results_btn.setEnabled(True)
    
    ###############################################################################
    # Enable radio specific proteciton levels
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Activate radio selection dropdown and initialize dictionary to store protection levels
    # when the radio specific level dropdown is checked.

    def radio_specific(self):
        self.radio_dropdown.setEnabled(self.radio_specific_levels.isChecked())
        self.radio_dropdown.clear()
        if self.radio_dropdown.isEnabled():
            emitapp.set_active_design(self.design_name_dropdown.currentText())
            radios = emitapp.modeler.components.get_radios()
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            for radio in radios:
                if radios[radio].has_rx_channels():
                    self.protection_levels[radio] = values
                    self.radio_dropdown.addItem(radio)
        else:
            self.radio_dropdown.clear()
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            self.protection_levels['Global'] = values
        self.global_protection_level = not self.radio_specific_levels.isChecked()
    
    ###############################################################################
    # Update legend table
    # ~~~~~~~~~~~~~~~~~~~
    # Update shown legend table values when the radio dropdown value changes. 

    def radio_dropdown_changed(self):
        if self.radio_dropdown.isEnabled():
            self.changing = True
            for row in range(self.protection_legend_table.rowCount()):
                item = self.protection_legend_table.item(row, 0)
                item.setText(str(self.protection_levels[self.radio_dropdown.currentText()][row]))
            self.changing = False

    ###############################################################################
    # Save legend table values
    # ~~~~~~~~~~~~~~~~~~~~~~~~
    # Save inputted radio protection level threshold values every time one is changed 
    # in the legend table.

    def table_changed(self):
        if self.changing == False:
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            if self.radio_dropdown.currentText() == '':
                index = 'Global'
            else:
                index = self.radio_dropdown.currentText()
            self.protection_levels[index] = values

    ###############################################################################
    # Save scenario matrix to as PNG file 
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Save the scenario matrix table as a PNG file.

    def save_image(self):
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
        else:
            table = self.interference_matrix

        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Save Scenario Matrix", "Scenario Matrix", "png (*.png)")
        image = QtGui.QImage(table.size(), QtGui.QImage.Format_ARGB32)
        table.render(image)
        image.save(fname[0])

    ###############################################################################
    # Save scenario matrix to Excel file 
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Write the scenario matrix results to an Excel file with color coding.

    def save_results_excel(self):
        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Save Scenario Matrix", "Protection Level Classification", "xlsx (*.xlsx)")
        
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
        else:
            table = self.interference_matrix

        if fname:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            header = self.tx_radios[:]
            header.insert(0, "Tx/Rx")
            worksheet.append(header)
            for row in range(2, table.rowCount()+2):
                worksheet.cell(row = row, column = 1, value = str(self.rx_radios[row-2]))
                for col in range(2, table.columnCount()+2):
                    text = str(table.item(row-2, col-2).text())
                    worksheet.cell(row = row, column = col, value = text)
                    cell = worksheet.cell(row, col)
                    cell.fill = PatternFill(start_color = self.color_dict[self.all_colors[col-2][row-2]][1][1:], 
                                            end_color = self.color_dict[self.all_colors[col-2][row-2]][1][1:], 
                                            fill_type = "solid")
            workbook.save(fname[0])

    ###############################################################################
    # Run interference type simulation 
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Run interference type simulation and classify results.

    def interference_results(self):
        # Initialize filter check marks and expected filter results
        self.interference_checks = [self.in_in_check.isChecked(), self.out_in_check.isChecked(),
                                  self.in_out_check.isChecked(), self.out_out_check.isChecked()]

        self.interference_filters =["TxFundamental:In-band", ["TxHarmonic/Spurious:In-band","Intermod:In-band", "Broadband:In-band"], 
                                    "TxFundamental:Out-of-band", ["TxHarmonic/Spurious:Out-of-band","Intermod:Out-of-band", "Broadband:Out-of-band"]]

        # Create list of problem types to analyze according to inputted filters
        filter = [i for (i,v) in zip(self.interference_filters, self.interference_checks) if v]

        if self.file_path_box.text() != "" and self.design_name_dropdown.currentText() != "":
            if self.previous_design != self.design_name_dropdown.currentText() or self.previous_project != self.file_path_box.text():
                self.previous_design = self.design_name_dropdown.currentText()
                self.previous_project = self.file_path_box.text()
                emitapp.set_active_design(self.design_name_dropdown.currentText())

                # Check if file is read-only
                if emitapp.save_project() == False:
                    msg = QtWidgets.QMessageBox()
                    msg.setWindowTitle("Writing Error")
                    msg.setText("An error occured while writing to the file. Is it readonly? Disk full? See AEDT log for more information.")
                    x = msg.exec()
                    return

                # Get results and radios
                self.rev = emitapp.results.analyze()
                self.tx_interferer = InterfererType().TRANSMITTERS
                self.rx_radios = self.rev.get_receiver_names()
                self.tx_radios = self.rev.get_interferer_names(self.tx_interferer)

                # Check if design is valid
                if self.tx_radios is None or self.rx_radios is None:
                    return

            # Classify the interference
            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~
            # Iterate over all the transmitters and receivers and compute the power
            # at the input to each receiver due to each of the transmitters. Compute
            # which, if any, type of interference occured.
            domain = emitapp.results.interaction_domain()
            self.all_colors, self.power_matrix = self.rev.interference_type_classification(domain, use_filter = True, filter_list = filter)

            # Save project and plot results on table widget
            emitapp.save_project()
            self.populate_table()

    ###############################################################################
    # Run protection level simulation 
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Run protection level simulation and classify results accroding to inputted 
    # threshold levels.

    def protection_results(self):
        # Initialize filter check marks and expected filter results
        self.protection_checks = [self.damage_check.isChecked(), self.overload_check.isChecked(),
                                  self.intermodulation_check.isChecked(), self.desensitization_check.isChecked()]

        self.protection_filters = ['damage', 'overload', 'intermodulation', 'desensitization']

        filter = [i for (i,v) in zip(self.protection_filters, self.protection_checks) if v]

        if self.file_path_box.text() != "" and self.design_name_dropdown.currentText() != "":
            if self.previous_design != self.design_name_dropdown.currentText() or self.previous_project != self.file_path_box.text():
                self.previous_design = self.design_name_dropdown.currentText()
                self.previous_project = self.file_path_box.text()
                emitapp.set_active_design(self.design_name_dropdown.currentText())

                # Check if file is read-only
                if emitapp.save_project() == False:
                    msg = QtWidgets.QMessageBox()
                    msg.setWindowTitle("Writing Error")
                    msg.setText("An error occured while writing to the file. Is it readonly? Disk full? See AEDT log for more information.")
                    x = msg.exec()
                    return
                
                # Get results and design radios
                self.tx_interferer = InterfererType().TRANSMITTERS
                self.rev = emitapp.results.analyze()
                self.rx_radios = self.rev.get_receiver_names()
                self.tx_radios = self.rev.get_interferer_names(self.tx_interferer)

            # Check if there are radios in the design
            if self.tx_radios is None or self.rx_radios is None:
                return

            domain = emitapp.results.interaction_domain()
            self.all_colors, self.power_matrix = self.rev.protection_level_classification(domain, 
                                                                                 self.global_protection_level, 
                                                                                 self.protection_levels['Global'], 
                                                                                 self.protection_levels, use_filter = True, 
                                                                                 filter_list = filter)

            self.populate_table()
    
    ###############################################################################
    # Populate the scenario matrix
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Create a scenario matrix view with the transmitters defined across the top
    # and receivers down the left-most column.

    def populate_table(self):
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
            button = self.protection_export_btn
            img_btn = self.protection_save_img_btn
        else:
            table = self.interference_matrix
            button = self.interference_export_btn
            img_btn = self.interference_save_img_btn

        num_cols = len(self.all_colors)
        num_rows = len(self.all_colors[0])
        table.setColumnCount(num_cols)
        table.setRowCount(num_rows)
        table.setVerticalHeaderLabels(self.rx_radios)
        table.setHorizontalHeaderLabels(self.tx_radios)

        for col in range(num_cols):
            for row in range(num_rows):
                item = QtWidgets.QTableWidgetItem(str(self.power_matrix[col][row]))
                table.setItem(row, col, item)
                cell = table.item(row, col)
                cell.setBackground(self.color_dict[self.all_colors[col][row]][0])
        
        button.setEnabled(True)
        img_btn.setEnabled(True)
    
    ###############################################################################
    # GUI closing event
    # ~~~~~~~~~~~~~~~~~
    # Close AEDT if the GUI is closed.
    def closeEvent(self, event):
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Closing GUI")
        msg.setText("Closing AEDT, please wait for GUI to close on its own.")
        x = msg.exec()
        emitapp.close_project()
        emitapp.close_desktop()

###############################################################################
# Run GUI
# ~~~~~~~
# Launch the GUI. If you want to run the GUI, uncomment the ``window.show()`` and
# ``app.exec_()`` method calls. 

if __name__ == '__main__' and  os.getenv("PYAEDT_DOC_GENERATION", "False") != "1":
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec()
else:
    emitapp.release_desktop(True,True)